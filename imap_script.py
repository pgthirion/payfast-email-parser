import imaplib
import email
from email.header import decode_header
from bs4 import BeautifulSoup
import openpyxl
import re
from datetime import datetime
import os

# ==============================
# CONFIGURATION
# ==============================
IMAP_SERVER = "" # Add your IMAP server here
EMAIL_ACCOUNT = "" # Add the email account that will receive the payfast emails here
PASSWORD = "" # Password for the email account
FOLDER = '"PayFast/Orders"'  # WorkMail exact folder name, with quotes

# ==============================
# CONNECT TO IMAP SERVER
# ==============================
mail = imaplib.IMAP4_SSL(IMAP_SERVER)
mail.login(EMAIL_ACCOUNT, PASSWORD)

status, select_data = mail.select(FOLDER)
if status != "OK":
    print(f"Could not select folder {FOLDER}: {select_data}")
    mail.logout()
    exit(1)

print("Folder selected successfully")

# Search for all unread emails
status, messages = mail.search(None, 'UNSEEN')
if status != "OK":
    print("Search failed:", messages)
    mail.logout()
    exit(1)

email_ids = messages[0].split()
print(f"Found {len(email_ids)} unread emails.\n")

# ==============================
# CREATE EXCEL FILE
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Payfast Orders"
ws.append(["Date", "Product", "Customer", "Discount", "Payment Method"])

# Match Sales formatting
date_format = "dd/mm/yyyy"
discount_number_format = "0.00"

# ==============================
# HELPER FUNCTIONS
# ==============================
def parse_price(price_str):
    """Convert R1 205 or R115 to float"""
    clean = re.sub(r"[^\d]", "", price_str)
    return float(clean) if clean else 0.0

def parse_email_date(date_str):
    """Parse email date string into timezone-free datetime"""
    try:
        dt = email.utils.parsedate_to_datetime(date_str)
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)  # remove timezone for Excel
        return dt
    except Exception:
        return None

# ===================================================
# BUNDLE NAMES (products that replace subproducts)
# ===================================================
BUNDLE_PRODUCTS = [
    "Afrikaans Huistaal Gr. 10 Oefenvraestelle 2022–2024 (Junie & November)",
    "Afrikaans Huistaal Gr. 11 Oefenvraestelle 2023 & 2024 (Junie & November)",
    "Afrikaans Huistaal Gr. 12 Oefenvraestelle 2023 & 2024 (Junie & September)",
    "Afrikaans EAT Gr. 10 Oefenvraestelle 2021–2024 (Junie & November)",
    "Afrikaans EAT Gr. 11 Oefenvraestelle 2023 & 2024 (Junie & November)",
    "Afrikaans EAT Gr. 12 Oefenvraestelle 2023 & 2024 (Junie & September)"
]

# ==============================
# PROCESS EACH EMAIL
# ==============================
for email_id in email_ids:
    status, data = mail.fetch(email_id, "(RFC822)")
    if status != "OK":
        print(f"Could not fetch email {email_id}")
        continue

    msg = email.message_from_bytes(data[0][1])

    # Get email date
    date_str = msg.get("Date", "")
    order_date = parse_email_date(date_str)

    # Get email body
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8")
                break
    else:
        body = msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8")

    soup = BeautifulSoup(body, "html.parser")

    # Customer
    customer_text = soup.find("p", string=re.compile(r"order from"))
    if customer_text:
        customer_match = re.search(r"order from (.+):", customer_text.get_text())
        customer = customer_match.group(1).strip() if customer_match else "Unknown"
    else:
        customer = "Unknown"

    # Products
    products = []
    for row in soup.find_all("tr", class_="order_item"):
        cols = row.find_all("td")
        if len(cols) >= 3:
            product_name = cols[0].get_text(strip=True)
            price_text = cols[2].get_text(strip=True)
            price = parse_price(price_text)
            products.append((product_name, price))

    # Subtotal
    subtotal_row = soup.find("tr", class_="order-totals-subtotal")
    subtotal = 0
    if subtotal_row:
        subtotal_text = subtotal_row.find_all("td")[-1].get_text(strip=True)
        subtotal = parse_price(subtotal_text)

    # Discount
    discount_row = soup.find("tr", class_="order-totals-discount")
    discount = 0
    if discount_row:
        discount_text = discount_row.find_all("td")[-1].get_text(strip=True)
        discount = parse_price(discount_text)

    discount_decimal = (discount / subtotal) if subtotal else 0

    # Payment method
    payment_method = "PF"

    # ==============================
    # STRIP SUBPRODUCTS IF BUNDLE FOUND
    # ==============================
    product_names = [p[0] for p in products]
    has_bundle = any(b for b in BUNDLE_PRODUCTS if b in product_names)

    if has_bundle:
        # Keep only the bundle product(s)
        products = [(p, pr) for p, pr in products if p in BUNDLE_PRODUCTS]

    # Append to Excel
    for product_name, price in products:
        ws.append([
            order_date if order_date else "",
            product_name,
            customer,
            discount_decimal,
            payment_method
        ])

    # Mark email as read
    mail.store(email_id, '+FLAGS', '\\Seen')

# ==============================
# APPLY FORMATTING
# ==============================
for cell in ws["A"]:  # Date column
    if cell.row != 1:
        cell.number_format = date_format

for cell in ws["D"]:  # Discount column
    if cell.row != 1:
        cell.number_format = discount_number_format

# ==============================
# SAVE EXCEL FILE
# ==============================
script_dir = os.path.dirname(os.path.abspath(__file__))
export_folder = os.path.join(script_dir, 'exports')
os.makedirs(export_folder, exist_ok=True)

current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
filename = f"PayfastOrders_{current_time}.xlsx"
file_path = os.path.join(export_folder, filename)

wb.save(file_path)
print(f"\nData saved to: {file_path}")

mail.logout()
